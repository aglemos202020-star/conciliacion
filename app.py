"""
Conciliador Bancario Web
========================
Correr con:  python app.py
Abrir en:    http://localhost:5000
"""

import os, re, json, base64, io, threading
from pathlib import Path
from datetime import date, datetime
from flask import Flask, request, jsonify, send_file, render_template_string
import anthropic, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ── CONFIG ──────────────────────────────────
API_KEY = os.environ.get("ANTHROPIC_API_KEY", "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")  # ← Podés poner tu clave acá o en Railway
UPLOAD_FOLDER = "uploads_tmp"
UMBRAL_MATCH  = 4
# ────────────────────────────────────────────

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB máx
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── Helpers ──────────────────────────────────

def find_col(keys, hints):
    lower = [k.lower().strip() for k in keys]
    for h in hints:
        for i, k in enumerate(lower):
            if h in k: return keys[i]
    return None

def normalize_monto(v):
    if v is None or v == "": return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    negativo = s.startswith("-")
    s = re.sub(r"[^\d,\.]", "", s)
    if not s: return 0.0
    if "," in s and "." in s: result = float(s.replace(".", "").replace(",", "."))
    elif "," in s: result = float(s.replace(",", "."))
    else: result = float(s) if s else 0.0
    return -result if negativo else result

def normalize_date(v):
    if not v: return ""
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})$", s)
    if m:
        y = ("20" + m.group(3)) if len(m.group(3)) == 2 else m.group(3)
        return f"{y}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    m = re.match(r"^(\d{4})[\/\-\.](\d{1,2})[\/\-\.](\d{1,2})$", s)
    if m: return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    if "T" in s: return s.split("T")[0]
    return s[:10]

def clean_cuit(c):
    return re.sub(r"[^\d]", "", str(c)) if c else ""

def extract_cuit_desc(desc):
    m = re.search(r"(?:^|\s)(\d{10,11})[\s\-]", str(desc))
    return m.group(1) if m else ""

def file_to_b64(path):
    with open(path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8")

def get_media_type(path):
    return {".pdf":"application/pdf",".jpg":"image/jpeg",
            ".jpeg":"image/jpeg",".png":"image/png"}.get(Path(path).suffix.lower(),"")

def extract_comprobante(path):
    client = anthropic.Anthropic(api_key=API_KEY)
    b64 = file_to_b64(path)
    mt  = get_media_type(path)
    is_img = Path(path).suffix.lower() in {".jpg",".jpeg",".png"}
    resp = client.messages.create(
        model="claude-opus-4-5", max_tokens=700,
        messages=[{"role":"user","content":[
            {"type":"image" if is_img else "document",
             "source":{"type":"base64","media_type":mt,"data":b64}},
            {"type":"text","text":(
                "Comprobante de transferencia bancaria argentina. "
                "Devolvé SOLO un JSON con: fecha (DD/MM/YYYY), monto (número sin símbolo), "
                "moneda, cuit_origen (solo dígitos), nombre_origen, cbu_origen, "
                "nombre_destino, cbu_destino, referencia, banco_origen. "
                "Null si no aparece. Sin backticks.")}
        ]}])
    text = "".join(b.text for b in resp.content if hasattr(b,"text"))
    try:
        return json.loads(re.sub(r"```json|```","",text).strip())
    except:
        return {"monto":None,"fecha":None,"error":"parse_error","raw":text}

def match_score(mov_fecha, mov_monto, mov_desc, comp):
    score, razones = 0, []
    c_monto = normalize_monto(comp.get("monto"))
    if c_monto and mov_monto:
        diff = abs(mov_monto - c_monto) / max(mov_monto, c_monto)
        if diff < 0.01: score += 3; razones.append("monto exacto")
        elif diff < 0.05: score += 1; razones.append("monto aprox")
    cuit_c = clean_cuit(comp.get("cuit_origen",""))
    cuit_d = extract_cuit_desc(mov_desc)
    if cuit_c and cuit_d and cuit_c[-10:] == cuit_d[-10:]:
        score += 3; razones.append("CUIT")
    c_fecha = normalize_date(comp.get("fecha",""))
    if mov_fecha and c_fecha:
        if mov_fecha == c_fecha: score += 2; razones.append("fecha exacta")
        else:
            try:
                if abs((date.fromisoformat(mov_fecha)-date.fromisoformat(c_fecha)).days)<=3:
                    score += 1; razones.append("fecha cercana")
            except: pass
    return score, razones

def conciliar(planilla_path, comp_paths):
    df = pd.read_excel(planilla_path, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    HINTS = {
        "fecha": ["fecha","date","día","dia"],
        "monto": ["haber","crédito","credito","importe","monto","ingreso","amount","entrada","credit"],
        "desc":  ["descripcion","descripción","concepto","detalle","movimiento","referencia","desc","glosa"],
        "tipo":  ["tipo","type","clase","category"],
    }
    cols = {k: find_col(df.columns.tolist(), v) for k,v in HINTS.items()}
    if not cols["monto"]:
        raise ValueError(f"No se encontró columna de monto. Columnas: {df.columns.tolist()}")

    mask = df[cols["monto"]].apply(normalize_monto) > 0
    if cols["desc"]:
        desc_up = df[cols["desc"]].fillna("").str.strip().str.upper()
        mask &= desc_up.str.startswith("TRANSFER")
        # Excluir CUITs propios o pagos que figuran como ingreso por error
        CUITS_EXCLUIDOS = ["30708635754", "30711747709"]  # La Santaniana, Terminal Pacheco
        for cuit_ex in CUITS_EXCLUIDOS:
            mask &= ~df[cols["desc"]].fillna("").str.contains(cuit_ex, na=False)
    ing = df[mask].copy()

    # Extraer comprobantes con IA
    comprobantes = []
    for p in comp_paths:
        try:
            data = extract_comprobante(p)
            comprobantes.append({"file": Path(p).name, "data": data})
        except Exception as e:
            comprobantes.append({"file": Path(p).name, "data": {"monto":None,"fecha":None}, "error":str(e)})

    # Cruzar
    rows = []
    for _, mov in ing.iterrows():
        fecha  = normalize_date(mov.get(cols["fecha"],"") if cols["fecha"] else "")
        monto  = normalize_monto(mov.get(cols["monto"],0))
        desc   = str(mov.get(cols["desc"],"") if cols["desc"] else "").strip()
        best_score, best_comp, best_razones = 0, None, []
        for c in comprobantes:
            s, r = match_score(fecha, monto, desc, c["data"])
            if s > best_score: best_score, best_comp, best_razones = s, c, r
        verificado = best_score >= UMBRAL_MATCH
        rows.append({
            "fecha":       fecha,
            "descripcion": desc,
            "importe":     round(monto, 2),
            "estado":      "Verificado" if verificado else "Sin comprobante",
            "comprobante": best_comp["file"] if verificado else "",
            "match":       " + ".join(best_razones) if verificado else "",
            "score":       best_score,
        })

    # Comprobantes que no matchearon con ningún movimiento
    comp_usados = {r["comprobante"] for r in rows if r["comprobante"]}
    sin_match = []
    for c in comprobantes:
        if c["file"] not in comp_usados:
            d = c.get("data", {})
            sin_match.append({
                "file":    c["file"],
                "fecha":   normalize_date(d.get("fecha", "")),
                "monto":   normalize_monto(d.get("monto") or 0),
                "nombre":  d.get("nombre_origen") or d.get("origen") or "—",
                "banco":   d.get("banco_origen") or "—",
                "cuit":    d.get("cuit_origen") or "—",
            })

    return {"rows": rows, "sin_match": sin_match}

# ── Exportadores ────────────────────────────

def export_excel(rows, sin_match=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliacion"
    headers = ["Fecha","Descripción","Importe","Estado","Comprobante","Criterios de match"]
    verde    = PatternFill("solid", fgColor="C6EFCE")
    amarillo = PatternFill("solid", fgColor="FFEB9C")
    hfill    = PatternFill("solid", fgColor="1F4E79")
    hfont    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    bfont    = Font(name="Arial", size=10)
    thin     = Side(style="thin", color="BFBFBF")
    border   = Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill=hfill; cell.font=hfont
        cell.alignment=Alignment(horizontal="center"); cell.border=border

    for r in rows:
        ws.append([r["fecha"],r["descripcion"],r["importe"],r["estado"],r["comprobante"],r["match"]])
        row_cells = ws[ws.max_row]
        fill = verde if r["estado"]=="Verificado" else amarillo
        for cell in row_cells:
            cell.fill=fill; cell.font=bfont; cell.border=border
            cell.alignment=Alignment(horizontal="left")

    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 55)
    ws.freeze_panes = "A2"

    # Hoja 2: comprobantes sin match
    if sin_match:
        ws2 = wb.create_sheet("Sin match en extracto")
        headers2 = ["Archivo", "Fecha", "Monto", "Remitente", "Banco", "CUIT"]
        ws2.append(headers2)
        for cell in ws2[1]:
            cell.fill=header_fill; cell.font=header_font
            cell.alignment=Alignment(horizontal="center"); cell.border=border
        rojo_fill = PatternFill("solid", fgColor="FFCCCC")
        for r in sin_match:
            ws2.append([r["file"], r["fecha"], r["monto"], r["nombre"], r["banco"], r["cuit"]])
            for cell in ws2[ws2.max_row]:
                cell.fill=rojo_fill; cell.font=body_font; cell.border=border
                cell.alignment=Alignment(horizontal="left")
        for col in ws2.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 50)
        ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

def export_pdf(rows, sin_match=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 fontSize=14, textColor=colors.HexColor("#1F4E79"),
                                 spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=9, textColor=colors.grey, spaceAfter=14)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=11)

    total = len(rows)
    verif = sum(1 for r in rows if r["estado"]=="Verificado")
    story = [
        Paragraph("Conciliación Bancaria", title_style),
        Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
                  f"Total ingresos: {total}  ·  Verificados: {verif}  ·  Sin comprobante: {total-verif}", sub_style),
    ]

    table_data = [["Fecha","Descripción","Importe","Estado","Comprobante","Match"]]
    for r in rows:
        table_data.append([
            r["fecha"],
            Paragraph(r["descripcion"][:60], cell_style),
            f"${r['importe']:,.0f}",
            r["estado"],
            Paragraph(r["comprobante"], cell_style),
            r["match"],
        ])

    col_widths = [2.2*cm, 8*cm, 2.8*cm, 3.2*cm, 4.5*cm, 5.5*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 8),
        ("FONTSIZE",    (0,1), (-1,-1), 8),
        ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID",        (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]
    # Colorear filas por estado
    for i, r in enumerate(rows, 1):
        if r["estado"] == "Verificado":
            style_cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#C6EFCE")))
        else:
            style_cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#FFEB9C")))
    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    # Tabla 2: comprobantes sin match
    if sin_match:
        story.append(Spacer(1, 1*cm))
        story.append(Paragraph("Comprobantes sin coincidencia en el extracto", title_style))
        story.append(Paragraph(f"Estos {len(sin_match)} comprobante(s) no se encontraron en la planilla de movimientos.", sub_style))
        t2_data = [["Archivo", "Fecha", "Monto", "Remitente", "Banco"]]
        for r in sin_match:
            monto_str = f"${r['monto']:,.0f}" if r['monto'] else "—"
            t2_data.append([
                r["file"],
                r["fecha"] or "—",
                monto_str,
                Paragraph(r["nombre"] or "—", cell_style),
                r["banco"] or "—",
            ])
        t2 = Table(t2_data, colWidths=[4*cm, 2.5*cm, 3*cm, 8*cm, 4*cm], repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#C0392B")),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("BACKGROUND",  (0,1), (-1,-1), colors.HexColor("#FFCCCC")),
            ("GRID",        (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 5),
        ]))
        story.append(t2)

    doc.build(story)
    buf.seek(0)
    return buf

# ── Rutas ────────────────────────────────────

HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Conciliador Bancario</title>
<style>
  :root{--verde:#C6EFCE;--verde-t:#1D9E75;--amarillo:#FFEB9C;--amarillo-t:#854F0B;
        --azul:#1F4E79;--bg:#F7F8FA;--card:#fff;--border:#E2E4E8;--text:#1A1A2E;--muted:#6B7280;}
  *{box-sizing:border-box;margin:0;padding:0;}
  body{font-family:Arial,sans-serif;background:var(--bg);color:var(--text);font-size:14px;}
  header{background:var(--azul);color:#fff;padding:1rem 2rem;display:flex;align-items:center;gap:12px;}
  header h1{font-size:18px;font-weight:600;}
  header span{font-size:13px;opacity:.7;}
  .container{max-width:1100px;margin:2rem auto;padding:0 1.5rem;}
  .card{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:1.5rem;margin-bottom:1.25rem;}
  .card-title{font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);margin-bottom:1rem;}
  .upload-grid{display:grid;grid-template-columns:1fr 1fr;gap:1rem;}
  .drop{border:1.5px dashed var(--border);border-radius:8px;padding:1.5rem;text-align:center;cursor:pointer;transition:.15s;}
  .drop:hover{background:#F0F4FF;border-color:#93C5FD;}
  .drop-icon{font-size:28px;margin-bottom:8px;}
  .drop-label{font-size:13px;color:var(--muted);margin-top:6px;}
  .file-list{margin-top:.75rem;display:flex;flex-direction:column;gap:6px;}
  .file-item{display:flex;align-items:center;gap:8px;font-size:12px;padding:5px 10px;
             background:#F9FAFB;border-radius:6px;border:1px solid var(--border);}
  .badge{font-size:10px;font-weight:700;padding:2px 6px;border-radius:4px;}
  .badge-xlsx{background:#EAF3DE;color:#3B6D11;}
  .badge-pdf{background:#FAECE7;color:#993C1D;}
  .badge-img{background:#E6F1FB;color:#185FA5;}
  .fname{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
  .rm{cursor:pointer;color:var(--muted);font-size:15px;line-height:1;}
  .rm:hover{color:#E24B4A;}
  .run-btn{width:100%;padding:.85rem;font-size:15px;font-weight:600;
           background:var(--azul);color:#fff;border:none;border-radius:8px;cursor:pointer;transition:.15s;}
  .run-btn:hover:not(:disabled){background:#163A5F;}
  .run-btn:disabled{opacity:.45;cursor:not-allowed;}
  .progress-wrap{margin-top:.75rem;}
  .progress-bar{height:5px;background:#E2E4E8;border-radius:3px;overflow:hidden;}
  .progress-fill{height:100%;background:var(--verde-t);border-radius:3px;transition:width .4s;}
  .status-msg{font-size:13px;color:var(--muted);text-align:center;margin-top:.5rem;min-height:18px;}
  .hidden{display:none!important;}
  .stats-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:1.25rem;}
  .stat{background:#F9FAFB;border:1px solid var(--border);border-radius:8px;padding:.85rem 1rem;}
  .stat-label{font-size:11px;color:var(--muted);margin-bottom:4px;}
  .stat-value{font-size:22px;font-weight:700;}
  .stat-value.green{color:var(--verde-t);}
  .stat-value.amber{color:var(--amarillo-t);}
  .results-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:.85rem;}
  .export-row{display:flex;gap:8px;}
  .btn-export{padding:.45rem 1rem;font-size:13px;font-weight:500;border:1px solid var(--border);
              background:#fff;border-radius:6px;cursor:pointer;transition:.15s;}
  .btn-export:hover{background:#F0F4FF;}
  .btn-xlsx{border-color:#3B6D11;color:#3B6D11;}
  .btn-pdf{border-color:#993C1D;color:#993C1D;}
  table{width:100%;border-collapse:collapse;font-size:13px;}
  th{text-align:left;padding:8px 10px;font-size:11px;font-weight:600;color:var(--muted);
     border-bottom:2px solid var(--border);white-space:nowrap;}
  td{padding:8px 10px;border-bottom:1px solid var(--border);vertical-align:middle;}
  tr:last-child td{border-bottom:none;}
  tr.verified td{background:var(--verde);}
  tr.missing td{background:var(--amarillo);}
  .chip{display:inline-flex;align-items:center;gap:4px;font-size:11px;font-weight:600;
        padding:3px 9px;border-radius:4px;}
  .chip-v{background:rgba(0,0,0,.08);color:#166534;}
  .chip-m{background:rgba(0,0,0,.08);color:var(--amarillo-t);}
  .dot{width:6px;height:6px;border-radius:50%;}
  .dot-g{background:var(--verde-t);}
  .dot-a{background:#EF9F27;}
  .comp-name{font-size:11px;color:#185FA5;margin-top:2px;}
  .match-tag{font-size:11px;color:var(--muted);}
  .error-box{background:#FEF2F2;border:1px solid #FCA5A5;border-radius:8px;
             padding:.85rem 1rem;margin-top:.75rem;font-size:13px;color:#991B1B;}
  .scroll-wrap{overflow-x:auto;}
  @media(max-width:700px){.upload-grid{grid-template-columns:1fr;}.stats-grid{grid-template-columns:1fr 1fr;}}
  .sinmatch-title{font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;
                  color:#991B1B;margin-bottom:.75rem;}
  tr.sinmatch td{background:#FFCCCC;}
  .chip-sm{display:inline-flex;align-items:center;gap:4px;font-size:11px;font-weight:600;
           padding:2px 8px;border-radius:4px;background:rgba(0,0,0,.08);color:#991B1B;}
</style>
</head>
<body>
<header>
  <div>
    <h1>Conciliador Bancario</h1>
    <span>Cruzá tus movimientos con los comprobantes de transferencia</span>
  </div>
</header>

<div class="container">
  <div class="card">
    <div class="card-title">Archivos</div>
    <div class="upload-grid">
      <div>
        <div class="drop" id="drop-xlsx" onclick="document.getElementById('inp-xlsx').click()">
          <div class="drop-icon">📊</div>
          <div style="font-weight:600">Planilla de movimientos</div>
          <div class="drop-label">Excel (.xlsx) del banco</div>
        </div>
        <input type="file" id="inp-xlsx" accept=".xlsx,.xls,.csv" style="display:none">
        <div class="file-list" id="list-xlsx"></div>
      </div>
      <div>
        <div class="drop" id="drop-comp" onclick="document.getElementById('inp-comp').click()">
          <div class="drop-icon">🗂</div>
          <div style="font-weight:600">Comprobantes de transferencia</div>
          <div class="drop-label">PDF, JPG o PNG — podés subir varios</div>
        </div>
        <input type="file" id="inp-comp" accept=".pdf,.jpg,.jpeg,.png" multiple style="display:none">
        <div class="file-list" id="list-comp"></div>
      </div>
    </div>
  </div>

  <button class="run-btn" id="run-btn" disabled onclick="runConciliation()">
    Conciliar transferencias
  </button>
  <div class="progress-wrap hidden" id="prog-wrap">
    <div class="progress-bar"><div class="progress-fill" id="prog-fill" style="width:0%"></div></div>
    <div class="status-msg" id="status-msg"></div>
  </div>
  <div id="error-area"></div>

  <div id="results-section" class="hidden" style="margin-top:1.5rem;">
    <div class="stats-grid" id="stats-grid"></div>
    <div class="card">
      <div class="results-header">
        <div class="card-title" style="margin-bottom:0">Resultados</div>
        <div class="export-row">
          <button class="btn-export btn-xlsx" onclick="exportFile('xlsx')">⬇ Excel</button>
          <button class="btn-export btn-pdf"  onclick="exportFile('pdf')">⬇ PDF</button>
        </div>
      </div>
      <div class="scroll-wrap">
        <table>
          <thead>
            <tr>
              <th style="width:90px">Fecha</th>
              <th>Descripción</th>
              <th style="width:120px;text-align:right">Importe</th>
              <th style="width:150px">Estado</th>
              <th>Comprobante / Criterios</th>
            </tr>
          </thead>
          <tbody id="results-body"></tbody>
        </table>
      </div>
    </div>
  <div id="sinmatch-section" class="hidden" style="margin-top:1rem;">
    <div class="card">
      <div class="results-header">
        <div class="sinmatch-title">⚠ Comprobantes sin coincidencia en el extracto</div>
        <div style="font-size:12px;color:var(--muted)" id="sinmatch-count"></div>
      </div>
      <div class="scroll-wrap">
        <table>
          <thead>
            <tr>
              <th style="width:160px">Archivo</th>
              <th style="width:90px">Fecha</th>
              <th style="width:120px;text-align:right">Monto</th>
              <th>Remitente</th>
              <th style="width:120px">Banco</th>
              <th style="width:130px">CUIT</th>
            </tr>
          </thead>
          <tbody id="sinmatch-body"></tbody>
        </table>
      </div>
    </div>
  </div>
  </div>
</div>

<script>
let xlsxFile = null, compFiles = [], lastResults = [], lastSinMatch = [];

function ext(n){ return (n||'').split('.').pop().toLowerCase(); }
function badgeClass(n){ const e=ext(n); if(['xlsx','xls','csv'].includes(e)) return 'badge-xlsx'; return e==='pdf'?'badge-pdf':'badge-img'; }
function badgeLabel(n){ const e=ext(n); if(e==='xlsx'||e==='xls') return 'XLSX'; if(e==='csv') return 'CSV'; if(e==='pdf') return 'PDF'; return 'IMG'; }

function renderXlsx(){
  document.getElementById('list-xlsx').innerHTML = xlsxFile
    ? `<div class="file-item"><span class="badge ${badgeClass(xlsxFile.name)}">${badgeLabel(xlsxFile.name)}</span>
       <span class="fname">${xlsxFile.name}</span>
       <span class="rm" onclick="xlsxFile=null;renderXlsx();check()">×</span></div>` : '';
}
function renderComp(){
  document.getElementById('list-comp').innerHTML = compFiles.map((f,i)=>
    `<div class="file-item"><span class="badge ${badgeClass(f.name)}">${badgeLabel(f.name)}</span>
     <span class="fname">${f.name}</span>
     <span class="rm" onclick="compFiles.splice(${i},1);renderComp();check()">×</span></div>`
  ).join('');
}
function check(){ document.getElementById('run-btn').disabled = !(xlsxFile && compFiles.length > 0); }

document.getElementById('inp-xlsx').onchange = e => { if(e.target.files[0]){xlsxFile=e.target.files[0];renderXlsx();check();} };
document.getElementById('inp-comp').onchange = e => { compFiles=[...compFiles,...Array.from(e.target.files)];renderComp();check(); };

['drop-xlsx','drop-comp'].forEach(id=>{
  const el=document.getElementById(id);
  el.addEventListener('dragover',e=>{e.preventDefault();el.style.background='#F0F4FF';});
  el.addEventListener('dragleave',()=>el.style.background='');
  el.addEventListener('drop',e=>{
    e.preventDefault();el.style.background='';
    const files=Array.from(e.dataTransfer.files);
    if(id==='drop-xlsx'){ const f=files.find(f=>['xlsx','xls','csv'].includes(ext(f.name))); if(f){xlsxFile=f;renderXlsx();check();}}
    else{ compFiles=[...compFiles,...files.filter(f=>['pdf','jpg','jpeg','png'].includes(ext(f.name)))];renderComp();check();}
  });
});

function setProgress(pct, msg){
  document.getElementById('prog-fill').style.width = pct+'%';
  document.getElementById('status-msg').textContent = msg;
}

async function runConciliation(){
  document.getElementById('error-area').innerHTML = '';
  document.getElementById('results-section').classList.add('hidden');
  document.getElementById('prog-wrap').classList.remove('hidden');
  document.getElementById('run-btn').disabled = true;
  setProgress(5, 'Subiendo archivos...');

  const fd = new FormData();
  fd.append('planilla', xlsxFile);
  compFiles.forEach(f => fd.append('comprobantes', f));

  try {
    setProgress(15, 'Procesando con IA — esto puede tardar 1-2 minutos...');
    const resp = await fetch('/conciliar', { method:'POST', body:fd });
    const data = await resp.json();
    if(data.error){ showError(data.error); return; }
    setProgress(100, '¡Conciliación completada!');
    lastResults  = data.rows;
    lastSinMatch = data.sin_match || [];
    renderResults(data.rows);
    renderSinMatch(data.sin_match || []);
  } catch(e) {
    showError('Error de conexión: ' + e.message);
  } finally {
    document.getElementById('run-btn').disabled = false;
  }
}

function showError(msg){
  document.getElementById('error-area').innerHTML = `<div class="error-box">${msg}</div>`;
  document.getElementById('prog-wrap').classList.add('hidden');
  document.getElementById('run-btn').disabled = false;
}

function renderResults(rows){
  const total   = rows.length;
  const verif   = rows.filter(r=>r.estado==='Verificado').length;
  const missing = total - verif;
  const totalMonto = rows.reduce((a,r)=>a+r.importe,0);

  document.getElementById('stats-grid').innerHTML = `
    <div class="stat"><div class="stat-label">Total ingresos</div><div class="stat-value">${total}</div></div>
    <div class="stat"><div class="stat-label">Verificados</div><div class="stat-value green">${verif}</div></div>
    <div class="stat"><div class="stat-label">Sin comprobante</div><div class="stat-value amber">${missing}</div></div>
    <div class="stat"><div class="stat-label">Total importe</div><div class="stat-value" style="font-size:16px">$${totalMonto.toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0})}</div></div>
  `;

  document.getElementById('results-body').innerHTML = rows.map(r => `
    <tr class="${r.estado==='Verificado'?'verified':'missing'}">
      <td style="color:#6B7280;font-variant-numeric:tabular-nums">${r.fecha||'—'}</td>
      <td>${r.descripcion||'—'}</td>
      <td style="text-align:right;font-weight:600;font-variant-numeric:tabular-nums">
        $${r.importe.toLocaleString('es-AR',{minimumFractionDigits:2,maximumFractionDigits:2})}
      </td>
      <td>
        ${r.estado==='Verificado'
          ? `<div class="chip chip-v"><div class="dot dot-g"></div>Verificado</div>`
          : `<div class="chip chip-m"><div class="dot dot-a"></div>Sin comprobante</div>`}
      </td>
      <td>
        ${r.comprobante ? `<div class="comp-name">${r.comprobante}</div>` : ''}
        ${r.match ? `<div class="match-tag">${r.match}</div>` : '—'}
      </td>
    </tr>
  `).join('');

  document.getElementById('results-section').classList.remove('hidden');
}

function renderSinMatch(items){
  const section = document.getElementById('sinmatch-section');
  if(!items || items.length === 0){ section.classList.add('hidden'); return; }
  document.getElementById('sinmatch-count').textContent = items.length + ' comprobante(s)';
  document.getElementById('sinmatch-body').innerHTML = items.map(r => `
    <tr class="sinmatch">
      <td><span class="chip-sm">📄</span> ${r.file}</td>
      <td style="color:var(--muted)">${r.fecha||'—'}</td>
      <td style="text-align:right;font-weight:600">
        ${r.monto ? '$'+r.monto.toLocaleString('es-AR',{minimumFractionDigits:0}) : '—'}
      </td>
      <td>${r.nombre||'—'}</td>
      <td>${r.banco||'—'}</td>
      <td style="font-family:monospace;font-size:12px">${r.cuit||'—'}</td>
    </tr>
  `).join('');
  section.classList.remove('hidden');
}

function exportFile(fmt){
  if(!lastResults.length) return;
  const fd = new FormData();
  fd.append('rows',      JSON.stringify(lastResults));
  fd.append('sin_match', JSON.stringify(lastSinMatch));
  fd.append('format', fmt);
  fetch('/exportar', { method:'POST', body:fd })
    .then(r=>r.blob())
    .then(blob=>{
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url;
      a.download = fmt==='xlsx' ? 'conciliacion.xlsx' : 'conciliacion.pdf';
      a.click();
      URL.revokeObjectURL(url);
    });
}
</script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/conciliar", methods=["POST"])
def conciliar_route():
    if API_KEY.startswith("sk-ant-..."):
        return jsonify({"error": "API key no configurada. Editá app.py y reemplazá API_KEY."})
    try:
        planilla = request.files["planilla"]
        comps    = request.files.getlist("comprobantes")

        plan_path = os.path.join(UPLOAD_FOLDER, "planilla" + Path(planilla.filename).suffix)
        planilla.save(plan_path)

        comp_paths = []
        for f in comps:
            p = os.path.join(UPLOAD_FOLDER, f.filename)
            f.save(p)
            comp_paths.append(p)

        result = conciliar(plan_path, comp_paths)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/exportar", methods=["POST"])
def exportar_route():
    rows      = json.loads(request.form["rows"])
    sin_match = json.loads(request.form.get("sin_match", "[]"))
    fmt       = request.form["format"]
    if fmt == "xlsx":
        buf = export_excel(rows, sin_match)
        return send_file(buf, download_name="conciliacion.xlsx",
                         as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        buf = export_pdf(rows, sin_match)
        return send_file(buf, download_name="conciliacion.pdf",
                         as_attachment=True, mimetype="application/pdf")

if __name__ == "__main__":
    print("="*50)
    print("  Conciliador Bancario Web")
    print("  Abrí tu navegador en: http://localhost:5000")
    print("="*50)
    app.run(debug=False, port=5000)
